import csv
import io
import json

import openpyxl

from text_as_data.export import results_to_csv_bytes, results_to_json_bytes, results_to_xlsx_bytes

SAMPLE_ROWS = [
    {"document_snippet": "About 200 people occupied...", "categoria": "protest", "justificativa": "clear demand"},
    {"document_snippet": "A music festival happened...", "categoria": "not_protest", "justificativa": "no claim"},
]


def test_results_to_csv_bytes_round_trips():
    content = results_to_csv_bytes(SAMPLE_ROWS)

    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    rows = list(reader)

    assert rows == SAMPLE_ROWS


def test_results_to_csv_bytes_handles_empty_list():
    assert results_to_csv_bytes([]) == b""


def test_results_to_csv_bytes_has_utf8_bom_for_excel():
    content = results_to_csv_bytes(SAMPLE_ROWS)

    assert content.startswith(b"\xef\xbb\xbf")


def test_results_to_csv_bytes_defuses_leading_formula_characters():
    rows = [{"justificativa": "=CMD|' /C calc.exe'!A0", "categoria": "x"}]

    content = results_to_csv_bytes(rows)

    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    row = next(reader)
    assert row["justificativa"] == "'=CMD|' /C calc.exe'!A0"


def test_results_to_csv_bytes_defuses_formula_prefixed_by_leading_whitespace():
    # Excel and LibreOffice both skip leading spaces/tabs when deciding
    # whether a cell is a formula -- a bare `.startswith()` check on the
    # raw value missed this, letting a value like " =cmd|...' " through
    # unquoted.
    rows = [{"justificativa": "  =CMD|' /C calc.exe'!A0", "categoria": "x"}]

    content = results_to_csv_bytes(rows)

    reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
    row = next(reader)
    assert row["justificativa"].startswith("'")


def test_results_to_xlsx_bytes_round_trips():
    content = results_to_xlsx_bytes(SAMPLE_ROWS)

    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header = next(rows_iter)
    rows = [dict(zip(header, values)) for values in rows_iter]

    assert rows == SAMPLE_ROWS


def test_results_to_xlsx_bytes_strips_illegal_control_characters():
    rows = [{"justificativa": "hello\x0bworld", "categoria": "x"}]

    content = results_to_xlsx_bytes(rows)  # must not raise IllegalCharacterError

    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    next(rows_iter)  # header
    data_row = next(rows_iter)
    assert data_row[0] == "helloworld"


def test_results_to_xlsx_bytes_defuses_leading_formula_characters():
    rows = [{"justificativa": "=SUM(A1:A10)", "categoria": "x"}]

    content = results_to_xlsx_bytes(rows)

    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    next(rows_iter)  # header
    data_row = next(rows_iter)
    assert data_row[0] == "'=SUM(A1:A10)"


def test_results_to_json_bytes_round_trips():
    content = results_to_json_bytes(SAMPLE_ROWS)

    assert json.loads(content.decode("utf-8")) == SAMPLE_ROWS
